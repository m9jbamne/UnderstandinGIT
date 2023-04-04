import openpyxl
import pytest


@pytest.mark.usefixtures("browser_invoke")
class Baseclass:

    @staticmethod
    def dataload(self):
        workbook = openpyxl.load_workbook("C:\\Users\\Rashmi\\E2EAutomation_022023\\Testdata.xlsx")
        sheet = workbook.active
        new_lst = []
        lst = []
        for i in range(1, sheet.max_row + 1):
            #print(i)
            for j in range(1, sheet.max_column + 1):
                t = sheet.cell(row=i, column=j).value
                lst.append(t)
                #print(lst)
            if lst not in new_lst:
                new_lst.append(tuple(lst))
                lst = []
        return new_lst
